from django_tables2.views import SingleTableView
from bootstrap_modal_forms.generic import (
    BSModalDeleteView, BSModalCreateView, BSModalUpdateView
)

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import (
    LoginRequiredMixin, PermissionRequiredMixin
)
from django.contrib.auth import logout
from django.contrib import messages
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.views.generic import DetailView, ListView
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.http import Http404, HttpRequest, HttpResponse, JsonResponse
from django.views.generic.edit import UpdateView

from .models import (
    Category, Course, CourseOccasion, CourseScheduleSegment, Block,
    Prerequisite, User, Profile,
    CategoryExam, CategoryCourse, AcademicYear, Exam, Report,
    PrivateCourse, ISPTemplate
)
from .tables import (
    ExamTable, ReportTable
)
from .filters import CourseFilter
from .forms import (
        CourseForm, BlockForm, ProfileForm, CourseOccasionForm,
        CourseScheduleSegmentForm, ExamForm,
        CategoryForm, ReportForm, PrivateCourseForm, UpdateUserForm,
        DeleteUserForm
)
from .rodatraden_modules.mixins import CorrectUserPermissionMixin
from .rodatraden_modules.functions import import_course_occasions, is_ajax

import openpyxl
import os
import math
import re
from io import BytesIO
from django.utils.crypto import get_random_string
from operator import itemgetter
from rodatraden import validator
from openpyxl.styles import Alignment, Font
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q, Manager
from django.core.exceptions import PermissionDenied
from django.core.mail import send_mail
from django.contrib.auth.models import Permission

def get_public_elective_course_occasions(block: Block):
    """Get public elective courses, that is courses not in the base block.
    Course occasions outside the base block start in year 3, period 4.
    """

    third_year = block.start_year + 2
    fourth_period_start_week = 30

    third_year_elective_course_occasions = block.courseoccasions.filter(
        academic_year__year=third_year,
        time_period__week__gte=fourth_period_start_week
    )

    course_occasions_last_two_years = block.courseoccasions.filter(
        academic_year__year__gt=third_year
    )

    elective_course_occasions = (third_year_elective_course_occasions
                                    | course_occasions_last_two_years)

    return elective_course_occasions


def index(request: HttpRequest):
    """Homepage of site."""

    context = {
            'latest_courses': Course.objects.all().order_by('-updated_at')[:7],
            'latest_blocks':
            Block.objects.filter(private=0).order_by('-updated_at')[:7],
            'categories': Category.objects.all().order_by('title'),
            'profiles': Profile.objects.all(),
            }

    return render(request, 'rodatraden/index.html', context)


def changelog(request: HttpRequest):
    """A site with changelogs."""

    return render(request, 'rodatraden/changelog.html')


def tools(request: HttpRequest):
    """Various tools for usage"""

    # Only staff can access this site
    if not request.user.is_staff:
        return redirect(reverse('index'))

    # Incoming info
    generate_results = None
    if request.method == 'POST':
        # Copying courseoccasions tool (legacy)
        if 'courseocc_copy' in request.POST:
            from_acyear = AcademicYear.objects.get(id=request.POST['from'])
            to_acyear = AcademicYear.objects.get(id=request.POST['to'])
            # Make sure that the years exist
            if from_acyear and to_acyear:
                # From all the current courseoccasions
                for courseocc in CourseOccasion.objects.filter(
                        academic_year=from_acyear
                        ):
                    # Only create a new for the new year if it does not already
                    # exist
                    if not CourseOccasion.objects.filter(
                            academic_year=to_acyear,
                            course=courseocc.course):
                        courseocc.pk = None
                        courseocc.academic_year = to_acyear
                        courseocc.slug = ''
                        courseocc.save()

        # Generate course occasions from scheduling rules
        if 'generate_occasions' in request.POST:
            generate_results = _generate_occasions_all_years(request)

    # Get all academic years
    acyears = AcademicYear.objects.all().order_by('year')

    context = {
            'acyears': acyears,
            'generate_results': generate_results,
    }

    return render(request, 'rodatraden/tools.html', context)


def sync_course_occasions(course, dry_run=True):
    """Compute and optionally apply scheduling changes for a single course.

    Compares the expected CourseOccasions (derived from the course's
    schedule segments) against what actually exists in the database.

    Returns a dict with keys:
        - created: list of dicts {year_title, period_title}
        - skipped_exists: list of dicts {year_title, period_title}
        - removed: list of dicts {year_title, period_title}
    """
    academic_years = AcademicYear.objects.all().order_by('year')
    created = []
    skipped_exists = []
    removed = []

    for academic_year in academic_years:
        year = academic_year.year
        segments = course.get_segments_for_year(year)

        # --- Create missing CourseOccasions ---
        for segment in segments:
            exists = CourseOccasion.objects.filter(
                course=course,
                academic_year=academic_year,
                time_period=segment.time_period,
            ).exists()

            entry = {
                'year_title': academic_year.title,
                'period_title': segment.time_period.title,
            }

            if exists:
                skipped_exists.append(entry)
                continue

            if not dry_run:
                CourseOccasion.objects.create(
                    course=course,
                    academic_year=academic_year,
                    time_period=segment.time_period,
                    weeks=segment.weeks,
                    official=True,
                    auto_generated=True,
                )
            created.append(entry)

        # --- Remove orphaned auto-generated CourseOccasions ---
        for co in CourseOccasion.objects.filter(
            course=course,
            academic_year=academic_year,
            auto_generated=True,
        ).select_related('time_period'):
            covered = any(
                seg.time_period_id == co.time_period_id
                for seg in segments
            )
            if not covered:
                removed.append({
                    'year_title': academic_year.title,
                    'period_title': co.time_period.title,
                })
                if not dry_run:
                    co.delete()

    return {
        'created': created,
        'skipped_exists': skipped_exists,
        'removed': removed,
    }


def _generate_occasions_all_years(request):
    """Generate CourseOccasion rows from scheduling segments for all years.

    Uses sync_course_occasions per course, then aggregates results by year.
    """
    dry_run = 'dry_run' not in request.POST  # checkbox unchecked = dry run

    # Collect results per year
    year_results = {}

    for course in Course.objects.filter(
        schedule_segments__isnull=False
    ).distinct().order_by('title'):
        result = sync_course_occasions(course, dry_run=dry_run)

        for entry in result['created']:
            yr = entry['year_title']
            year_results.setdefault(yr, {'created': [], 'skipped_exists': [], 'removed': []})
            year_results[yr]['created'].append(
                f'{course.title} — {entry["period_title"]}')

        for entry in result['skipped_exists']:
            yr = entry['year_title']
            year_results.setdefault(yr, {'created': [], 'skipped_exists': [], 'removed': []})
            year_results[yr]['skipped_exists'].append(
                f'{course.title} — {entry["period_title"]}')

        for entry in result['removed']:
            yr = entry['year_title']
            year_results.setdefault(yr, {'created': [], 'skipped_exists': [], 'removed': []})
            year_results[yr]['removed'].append(
                f'{course.title} — {entry["period_title"]}')

    # Also check for orphaned auto_generated occasions from courses with no segments
    for co in CourseOccasion.objects.filter(
        auto_generated=True,
    ).exclude(
        course__schedule_segments__isnull=False,
    ).select_related('course', 'time_period', 'academic_year'):
        yr = co.academic_year.title
        year_results.setdefault(yr, {'created': [], 'skipped_exists': [], 'removed': []})
        year_results[yr]['removed'].append(
            f'{co.course.title} — {co.time_period.title}')
        if not dry_run:
            co.delete()

    # Convert to list sorted by year title
    results = []
    for yr in sorted(year_results.keys()):
        data = year_results[yr]
        if data['created'] or data['skipped_exists'] or data['removed']:
            results.append({
                'year': yr,
                'dry_run': dry_run,
                'created': data['created'],
                'skipped_exists': data['skipped_exists'],
                'removed': data['removed'],
            })

    return results


##########
# ERRORS #
##########


def rt400(request: HttpRequest, exception=None):
    return render(request, 'rodatraden/400.html')


def rt403(request: HttpRequest, exception=None):
    return render(request, 'rodatraden/403.html')


def rt404(request: HttpRequest, exception=None):
    return render(request, 'rodatraden/404.html')


def rt500(request: HttpRequest, exception=None):
    return render(request, 'rodatraden/500.html')

#########
# USERS #
#########

class UserUpdate(CorrectUserPermissionMixin, UpdateView):
    """Update view for user profile information."""

    model = User
    form_class = UpdateUserForm
    # Check against username since users don't have slugs
    template_name = 'rodatraden/user/update_user_form.html'
    success_url = reverse_lazy('index')


@login_required
def user_change_password(request: HttpRequest, username: str, pk: int):
    """View for changing the logged-in user's password."""

    # Only allow the user to change their own password.
    # Note: pk is already an int from Django's <int:pk> URL converter.
    if request.user.username != username or request.user.id != pk:
        return redirect(reverse('index'))

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Keep the user logged in after password change
            update_session_auth_hash(request, user)
            return redirect(reverse('user-update', kwargs={
                'username': username, 'pk': pk
            }))
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'rodatraden/user/change_password.html', {
        'form': form
    })


@login_required
def user_delete(request: HttpRequest, username, pk):

    if request.user.username != username and request.user.id != pk:
        return redirect(reverse('index'))

    if request.method == 'POST':
        form = DeleteUserForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            if data['email'] == request.user.email:
                user = User.objects.get(pk=request.user.id)
                logout(request)
                user.delete()
                return redirect(reverse('index'))
    else:
        form = DeleteUserForm()

    return render(request, 'rodatraden/user/delete_user_form.html', {'form': form})

#############
## REPORTS ##
#############

class ReportCreate(BSModalCreateView):
    """Creation view for reports."""

    model = Report
    form_class = ReportForm
    template_name = 'rodatraden/report/report_create.html'
    success_message = 'Tack för din hjälp!'

    def get_initial(self):
        initial = super().get_initial()
        if self.request.user.is_authenticated:
            initial['from_email'] = self.request.user.email
        return initial

    def form_valid(self, form):
        response = super().form_valid(form)

        # Send email notification to users who can view reports, but only on
        # the actual form submission (not the AJAX validation request).
        if not is_ajax(self.request):
            self._send_report_notification(self.object)

        return response

    def _send_report_notification(self, report):
        """Send an email notification about the new report to all users who
        have the 'receive_report_email' permission (directly or via a group).
        Superusers are NOT automatically included; assign the permission
        explicitly to each user or group that should receive the emails."""

        try:
            perm = Permission.objects.get(
                codename='receive_report_email',
                content_type__app_label='rodatraden'
            )
        except Permission.DoesNotExist:
            return

        # Find all active users with the permission directly or via a group.
        users_with_perm = User.objects.filter(
            Q(user_permissions=perm) |
            Q(groups__permissions=perm),
            is_active=True,
        ).distinct()

        recipient_emails = [
            user.email for user in users_with_perm if user.email
        ]

        if recipient_emails:
            send_mail(
                subject=f'Ny rapport: {report.subject}',
                message=(
                    f'En ny rapport har skapats på Röda Tråden.\n\n'
                    f'Ämne: {report.subject}\n'
                    f'Från: {report.from_email}\n\n'
                    f'Meddelande:\n{report.message}\n\n'
                    f'Logga in för att hantera rapporten.'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=recipient_emails,
                fail_silently=True,
            )

    def get_success_url(self):
        # Return to last page
        return self.request.META['HTTP_REFERER']


class ReportList(LoginRequiredMixin, PermissionRequiredMixin, SingleTableView):
    """List view for reports."""

    permission_required = 'rodatraden.view_report'
    model = Report
    table_class = ReportTable
    template_name = 'rodatraden/report/report_list.html'


class ReportUpdate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalUpdateView):
    """Update view for reports."""

    model = Report
    form_class = ReportForm
    permission_required = 'rodatraden.change_report'
    template_name = 'rodatraden/report/report_update.html'
    success_message = 'Rapport uppdaterades utan problem'

    def get_success_url(self):
        # Return to last page
        return self.request.META['HTTP_REFERER']


class ReportDelete(LoginRequiredMixin, PermissionRequiredMixin,
                   BSModalDeleteView):
    """Delete view for reports."""

    model = Report
    permission_required = 'rodatraden.delete_report'
    template_name = 'rodatraden/report/report_confirm_delete.html'
    success_message = 'Rapport raderad'

    def get_success_url(self):
        # Return to last page
        return reverse_lazy('report-list')

###########
## EXAMS ##
###########

class ExamList(SingleTableView):
    """List view for exams"""

    model = Exam
    table_class = ExamTable
    template_name = 'rodatraden/exam/exam_list.html'


class ExamDetail(DetailView):
    """Detail view for exams"""

    model = Exam
    template_name = 'rodatraden/exam/exam_detail.html'


class ExamCreate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalCreateView):
    """Creation view for exams."""

    permission_required = 'rodatraden.add_exam'
    model = Exam
    form_class = ExamForm
    template_name = 'rodatraden/exam/exam_create.html'
    success_message = 'Examina skapades utan problem'
    success_url = reverse_lazy('exam-list')


class ExamUpdate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalUpdateView):
    """Update view for exams."""

    permission_required = 'rodatraden.change_exam'
    model = Exam
    form_class = ExamForm
    template_name = 'rodatraden/exam/exam_update.html'
    success_message = 'Examina uppdaterades utan problem'

    def get_success_url(self):
        # Return to last page
        return self.request.META['HTTP_REFERER']


class ExamDelete(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalDeleteView):
    """Delete view for exams."""

    permission_required = 'rodatraden.delete_exam'
    model = Exam
    template_name = 'rodatraden/exam/exam_confirm_delete.html'
    success_message = 'Examina togs bort utan problem'
    success_url = reverse_lazy('exam-list')

###########
# COURSES #
###########

def course_list(request: HttpRequest):
    """Custom list view for courses that both filters and sorts the courses."""


    def prepare_page_path(request: HttpRequest):
        """Tweaks the path so that it ends with "page=". This is so that the
        page number later can be appended to complete the path. Makes sure the
        query string doesn't contain duplicate path parameters."""

        path = request.path
        full_path = request.get_full_path()

        # We modify the existing query string if it exists.
        if '?' in full_path:
            # Isolate the GET-queries.
            query_string = full_path.split('?')[1]
            queries = query_string.split('&')

            # Remove any pre-existing page query.
            queries = [query for query in queries if not 'page=' in query]

            # Append unfinished query.
            queries.append('page=')

            # Reassemble the string and append it.
            query_string = '?' + '&'.join(queries)
            path += query_string

        # Otherwise we append a new one.
        else:
            path += '?page='
    
        return path


    courses_per_page = 15

    filter = CourseFilter(request.GET)
    paginator = Paginator(filter.qs, courses_per_page)
    paginator.ELLIPSIS = '…'

    # Default to the first page (1-based index).
    page_index = request.GET.get('page', 1)
    page = paginator.get_page(page_index)

    page_numbers = paginator.get_elided_page_range(page_index, on_each_side=2, on_ends=1)

    # Header information needed by the template that the course list template
    # extends.
    table_header_items = [
        {
            'name': 'title',
            'verbose_name': 'Kursnamn',
            'is_sortable': True
        },
        {
            'name': 'ects',
            'verbose_name': 'Poäng',
            'is_sortable': True
        },
        {
            'name': 'level',
            'verbose_name': 'Nivå',
            'is_sortable': True
        },
        {
            'name': 'categories',
            'verbose_name': 'Kategorier',
            'is_sortable': False
        }
    ]

    context = {
        'filter': filter,
        # The page is renamed because the parent template expects a list of
        # courses named 'courses'.
        'courses': page,
        'page_numbers': page_numbers,
        'ellipsis': paginator.ELLIPSIS,
        'unfinished_page_path': prepare_page_path(request),
        'table_header_items': table_header_items
    }

    return render(request, 'rodatraden/course/course_list.html', context)


class CourseDetail(DetailView):
    """Detail view for courses."""

    model = Course
    template_name = 'rodatraden/course/course_detail.html'

    def get_context_data(self, **kwargs):
        """Extend get_context_data to return courses that is required for the
        given course."""

        context = super().get_context_data(**kwargs)

        # Get all prerequisites that this course is appart of, meaning that the
        # queryset will inform which courses require this course.
        prerequisites: Manager[Prerequisite] = self.object.equivalent_prerequisites.all()

        courses_requiring_this_course = [
            prerequisite.course for prerequisite in prerequisites]

        # Keep only distinct courses.
        courses_requiring_this_course = list(set(courses_requiring_this_course))
        courses_requiring_this_course.sort(key=lambda course: course.title)

        context['courses_requiring_this_course'] = courses_requiring_this_course

        return context


class CourseCreate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalCreateView):
    """Creation view for courses."""

    permission_required = 'rodatraden.add_course'
    model = Course
    form_class = CourseForm
    template_name = 'rodatraden/course/course_create.html'
    success_message = 'Kursen skapades utan problem'
    success_url = reverse_lazy('course-list')


class CourseUpdate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalUpdateView):
    """Update view for courses."""

    model = Course
    form_class = CourseForm
    permission_required = 'rodatraden.change_course'
    template_name = 'rodatraden/course/course_update.html'
    success_message = 'Kursen uppdaterades utan problem'

    def get_success_url(self):
        # Return to last page
        return self.request.META['HTTP_REFERER']


class CourseDelete(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalDeleteView):
    """Delete view for courses."""

    permission_required = 'rodatraden.delete_course'
    model = Course
    template_name = 'rodatraden/course/course_confirm_delete.html'
    success_message = 'Kursen togs bort utan problem'
    success_url = reverse_lazy('course-list')

############################
# COURSE SCHEDULE SEGMENTS #
############################

class SegmentCreate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalCreateView):
    """Creation view for schedule segments."""

    permission_required = 'rodatraden.can_manage_scheduling'
    model = CourseScheduleSegment
    form_class = CourseScheduleSegmentForm
    template_name = 'rodatraden/segment/segment_create.html'
    success_message = 'Schemaläggningssegment skapat'

    def get_success_url(self):
        return reverse('course-detail', kwargs={
            'slug': self.object.course.slug
        })


class SegmentUpdate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalUpdateView):
    """Update view for schedule segments."""

    permission_required = 'rodatraden.can_manage_scheduling'
    model = CourseScheduleSegment
    form_class = CourseScheduleSegmentForm
    template_name = 'rodatraden/segment/segment_update.html'
    success_message = 'Schemaläggningssegment uppdaterat'

    def get_success_url(self):
        return reverse('course-detail', kwargs={
            'slug': self.object.course.slug
        })


class SegmentDelete(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalDeleteView):
    """Delete view for schedule segments."""

    permission_required = 'rodatraden.can_manage_scheduling'
    model = CourseScheduleSegment
    template_name = 'rodatraden/segment/segment_confirm_delete.html'
    success_message = 'Schemaläggningssegment raderat'

    def get_success_url(self):
        return reverse('course-detail', kwargs={
            'slug': self.object.course.slug
        })


@login_required
def segment_execute(request, pk):
    """Preview or apply scheduling sync for a single course.

    GET  → returns the preview modal with dry-run results.
    POST → applies changes and redirects back to the course detail page.
    """
    course = get_object_or_404(Course, pk=pk)

    if not request.user.has_perm('rodatraden.can_manage_scheduling'):
        raise PermissionDenied

    if request.method == 'POST':
        result = sync_course_occasions(course, dry_run=False)
        n_created = len(result['created'])
        n_removed = len(result['removed'])
        parts = []
        if n_created:
            parts.append(f'{n_created} kurstillfällen skapades')
        if n_removed:
            parts.append(f'{n_removed} kurstillfällen togs bort')
        msg = ', '.join(parts) if parts else 'Inga ändringar behövdes'
        messages.success(request, f'Schemaläggning verkställd: {msg}.')
        return redirect(reverse('course-detail', kwargs={'slug': course.slug}))

    # GET — dry-run preview
    result = sync_course_occasions(course, dry_run=True)
    context = {
        'course': course,
        'result': result,
    }
    return render(request, 'rodatraden/segment/segment_execute.html', context)


###################
# COURSEOCCASIONS #
###################


class CourseOccasionDetail(DetailView):
    """Detail view for courseoccasions."""

    model = CourseOccasion
    template_name = 'rodatraden/courseoccasion/courseoccasion_detail.html'

    def get_queryset(self):
        """Extend get_queryset to filter also by year"""

        self.qs = super().get_queryset()

        return self.qs.filter(
                academic_year__year=self.kwargs['year'])

    def get_context_data(self, **kwargs):
        """Extend get_context_data to return courses that is required for the
        given course."""

        context = super().get_context_data(**kwargs)

        # Get all prerequisites that this course is appart of, meaning that the
        # queryset will inform which courses require this course.
        prerequisites: Manager[Prerequisite] = self.object.course.equivalent_prerequisites.all()

        courses_requiring_this_course = [
            prerequisite.course for prerequisite in prerequisites]

        # Keep only distinct courses.
        courses_requiring_this_course = list(set(courses_requiring_this_course))
        courses_requiring_this_course.sort(key=lambda course: course.title)

        context['courses_requiring_this_course'] = courses_requiring_this_course

        return context


def courseoccasion_info(request: HttpRequest):
    """Small info view for courseoccasions used in blocks."""

    year = int(request.GET.get('year', ''))
    slug = request.GET.get('slug', '')

    # Get valid IDs of the unmet prerequisites.
    unmet_prerequisite_ids = [
        int(id) for id in request.GET.getlist('unmet[]') if id.isdigit()]

    unmet_prerequisites = Prerequisite.objects.filter(id__in=unmet_prerequisite_ids)

    courseoccasion = get_object_or_404(CourseOccasion, academic_year__year=year,
        slug=slug)

    context = {
        'courseoccasion': courseoccasion,
        'course': courseoccasion.course,
        'unmet_prerequisites': unmet_prerequisites,
    }

    return render(request, 'rodatraden/courseoccasion/courseoccasion_info.html',
        context)


class CourseOccasionCreate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalCreateView):
    """Creation view for courseoccasions."""

    permission_required = 'rodatraden.add_courseoccasion'
    model = CourseOccasion
    form_class = CourseOccasionForm
    template_name = 'rodatraden/courseoccasion/courseoccasion_create.html'
    success_message = 'Kurstillfället skapades utan problem'
    #success_url = reverse_lazy('courseoccasion-list')

    def get_success_url(self):
        # Return to last page
        return self.request.META['HTTP_REFERER']


class CourseOccasionUpdate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalUpdateView):
    """Update view for courseoccasions."""

    model = CourseOccasion
    form_class = CourseOccasionForm
    permission_required = 'rodatraden.change_courseoccasion'
    template_name = 'rodatraden/courseoccasion/courseoccasion_update.html'
    success_message = 'Kurstillfälle uppdaterat'

    def get_success_url(self):
        # Return to last page
        return self.request.META['HTTP_REFERER']


class CourseOccasionDelete(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalDeleteView):
    """Delete view for courseoccasions."""

    model = CourseOccasion
    permission_required = 'rodatraden.delete_courseoccasion'
    template_name = 'rodatraden/courseoccasion/courseoccasion_confirm_delete.html'
    success_message = 'Kurstillfället togs bort utan problem'

    def get_success_url(self):
        return reverse_lazy('courseoccasion-list')

############
# PROFILES #
############

class ProfileList(ListView):
    """List view for profiles."""

    model = Profile
    template_name = 'rodatraden/profile/profile_list.html'


class ProfileDetail(DetailView):
    """Detail view for profiles."""

    model = Profile
    template_name = 'rodatraden/profile/profile_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tracks'] = self.object.track_set.all()

        # Queries to be used below.
        included_courses_query = Q(tracks__profile__id=self.object.id)
        core_courses_query = Q(core_belonging__profile__id=self.object.id)
        related_courses_query = (included_courses_query | core_courses_query)

        # Courses related to the profile (included courses and core courses).
        included_courses = Course.objects.filter(included_courses_query)
        core_courses = Course.objects.filter(core_courses_query)
        related_courses = Course.objects.filter(related_courses_query).distinct().order_by('title')

        profile_category = Category.objects.get(title='Profilkurs')

        # Attach profile and core labels to each course.
        labeled_courses = []
        for course in related_courses:
            labeled_courses.append({
                'course': course,
                'is_profile': profile_category in course.categories.all(),
                'is_core': course in core_courses
            })

        context['labeled_courses'] = labeled_courses
        context['profile_id'] = self.object.id

        # Get public blocks related to the profile.
        blocks = Block.objects.filter(track__profile__id=self.object.id,
                                           private=False)

        # Put partial block info for related blocks in a JSON format so that it
        # is usable in JavaScript. This will be used to render all the block
        # schedules of a profile.
        blocks_json = []
        for block in blocks:
            if self.object.block_schedule_view_setting == self.object.BlockScheduleViewSetting.extended:
                elective_course_occasions = get_public_elective_course_occasions(block)
                course_occasions_json = [occasion.as_json() for occasion in elective_course_occasions]
            else:
                course_occasions_json = [course.as_json() for course in block.courseoccasions.all()]

            blocks_json.append({
                'title': block.title,
                'startYear': block.start_year,
                'courseOccasions': course_occasions_json
            })

        context['blocks'] = blocks_json

        return context


class ProfileCreate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalCreateView):
    """Creation view for profiles."""

    permission_required = 'rodatraden.add_profile'
    model = Profile
    form_class = ProfileForm
    template_name = 'rodatraden/profile/profile_create.html'
    success_message = 'Profilen skapades utan problem'
    success_url = reverse_lazy('profile-list')


class ProfileUpdate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalUpdateView):
    """Update view for profiles."""

    permission_required = 'rodatraden.change_profile'
    model = Profile
    form_class = ProfileForm
    template_name = 'rodatraden/profile/profile_update.html'
    success_message = 'Profilen uppdaterades utan problem'

    def get_success_url(self):
        # Return to last page
        return self.request.META['HTTP_REFERER']


class ProfileDelete(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalDeleteView):
    """Delete view for profiles."""

    permission_required = 'rodatraden.delete_profile'
    model = Profile
    template_name = 'rodatraden/profile/profile_confirm_delete.html'
    success_message = 'Profilen togs bort utan problem'
    success_url = reverse_lazy('profile-list')

##############
# CATEGORIES #
##############

class CategoryList(ListView):
    """List view for categories."""

    model = Category
    template_name = 'rodatraden/category/category_list.html'


class CategoryDetail(DetailView):
    """Detail view for categories."""

    model = Category
    template_name = 'rodatraden/category/category_detail.html'

    def get_context_data(self, **kwargs):
        """Extend get_context_data to return the connections between categories
        and course"""

        context = super().get_context_data(**kwargs)
        context['courses'] = CategoryCourse.objects.filter(
            category=context['category']
        )

        return context


class CategoryCreate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalCreateView):
    """Creation view for categories."""

    permission_required = 'rodatraden.add_category'
    model = Category
    form_class = CategoryForm
    template_name = 'rodatraden/category/category_create.html'
    success_message = 'Kategorin skapades utan problem'
    success_url = reverse_lazy('category-list')


class CategoryUpdate(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalUpdateView):
    """Update view for categories."""

    permission_required = 'rodatraden.change_category'
    model = Category
    form_class = CategoryForm
    template_name = 'rodatraden/category/category_update.html'
    success_message = 'Kategori uppdaterades utan problem'

    def get_success_url(self):
        # Return to last page
        return self.request.META['HTTP_REFERER']


class CategoryDelete(LoginRequiredMixin, PermissionRequiredMixin,
        BSModalDeleteView):
    """Deletion view for categories"""

    permission_required = 'rodatraden.delete_category'
    model = Category
    template_name = 'rodatraden/category/category_confirm_delete.html'
    success_message = 'Kategori togs bort utan problem'
    success_url = reverse_lazy('category-list')

###################
# PRIVATE COURSES #
###################

class PrivateCourseList(CorrectUserPermissionMixin, LoginRequiredMixin,
        ListView):
    """List view for private courses."""

    model = PrivateCourse
    template_name = 'rodatraden/privatecourse/privatecourse_list.html'
    # The parent template expects a list of courses named 'courses'.
    context_object_name = 'courses'

    def get_queryset(self):
        """Extend get_queryset to filter only private courses for user."""

        qs = super().get_queryset()

        sort_order = self.request.GET.get('sort_order', 'title')

        return qs.filter(
            user__username=self.kwargs['username']
        ).order_by(sort_order)


    def get_context_data(self):
        # Header information needed by the template that the course list template
        # extends.
        table_header_items = [
            {
                'name': 'title',
                'verbose_name': 'Kursnamn',
                'is_sortable': True
            },
            {
                'name': 'ects',
                'verbose_name': 'Poäng',
                'is_sortable': True
            },
            {
                'name': 'year',
                'verbose_name': 'Läsår',
                'is_sortable': True
            },
            {
                'name': 'start',
                'verbose_name': 'Start',
                'is_sortable': True
            },
            {
                'name': 'weeks',
                'verbose_name': 'Veckor',
                'is_sortable': True
            }
        ]

        context = super().get_context_data()
        context['table_header_items'] = table_header_items
        return context


class PrivateCourseDetail(CorrectUserPermissionMixin, LoginRequiredMixin,
        DetailView):
    """Detail view for private courses."""

    model = PrivateCourse
    template_name = 'rodatraden/privatecourse/privatecourse_detail.html'


class PrivateCourseUpdate(CorrectUserPermissionMixin, LoginRequiredMixin,
        BSModalUpdateView):
    """Update view for private courses"""

    model = PrivateCourse
    form_class = PrivateCourseForm
    template_name = 'rodatraden/privatecourse/privatecourse_update.html'
    success_message = 'Ändringarna i blockschemat sparades'

    def get_success_url(self):
        # Return to last page
        return self.request.META['HTTP_REFERER']


class PrivateCourseCreate(CorrectUserPermissionMixin, LoginRequiredMixin,
        BSModalCreateView):
    """Create view for private courses."""

    model = PrivateCourse
    form_class = PrivateCourseForm
    template_name = 'rodatraden/privatecourse/privatecourse_create.html'
    success_message = 'Egen kurs skapad'

    def get_success_url(self, **kwargs):
        return reverse_lazy('privatecourse-list',
                kwargs={'username':self.kwargs['username']})


class PrivateCourseDelete(CorrectUserPermissionMixin, LoginRequiredMixin,
        BSModalDeleteView):
    """Delete view for private courses."""

    model = PrivateCourse
    template_name = 'rodatraden/privatecourse/privatecourse_confirm_delete.html'
    success_message = 'Egen kurs raderat'

    def get_success_url(self, **kwargs):
        return reverse_lazy('privatecourse-list',
                kwargs={'username':self.kwargs['username']})

##########
# BLOCKS #
##########

class BlockList(CorrectUserPermissionMixin, LoginRequiredMixin, ListView):
    """List view for blocks."""

    model = Block
    template_name = 'rodatraden/block/block_list.html'

    def get_queryset(self, *args, **kwargs):
        self.qs = super().get_queryset()
        # Filter the blocks for the given user
        qs = self.qs.filter(user__username=self.kwargs['username']).order_by('title')
        return qs


class BlockUpdate(CorrectUserPermissionMixin, LoginRequiredMixin,
                  BSModalUpdateView):
    """Update view for blocks."""

    model = Block
    form_class = BlockForm
    template_name = 'rodatraden/block/block_update.html'
    success_message = 'Ändringarna i blockschemat sparades'

    def get_success_url(self):
        old_slug = self.kwargs['slug']
        new_slug = self.object.slug

        if new_slug == old_slug:
            # Return to last page
            return self.request.META['HTTP_REFERER']
        else:
            # Return to the updated url since the user renamed the block
            return reverse_lazy('block-detail',
                                kwargs={'username': self.kwargs['username'], 'slug': new_slug})


class BlockCreate(CorrectUserPermissionMixin, LoginRequiredMixin,
        BSModalCreateView):
    """Create view for blocks."""

    model = Block
    form_class = BlockForm
    template_name = 'rodatraden/block/block_create.html'
    success_message = 'Blockschema skapat'

    def form_valid(self, form):
        """Extend form_valid to add current user to block."""

        instance = form.save(commit=False)
        instance.user = self.request.user

        return super(BlockCreate, self).form_valid(form)

    def get_success_url(self, **kwargs):
        return reverse_lazy('block-list',
                            kwargs={'username':self.kwargs['username']})


class BlockDelete(CorrectUserPermissionMixin, LoginRequiredMixin,
        BSModalDeleteView):
    """Delete view for blocks."""

    model = Block
    template_name = 'rodatraden/block/block_confirm_delete.html'
    success_message = 'Blockschema raderat'

    def get_success_url(self, **kwargs):
        return reverse_lazy('block-list',
                            kwargs={'username':self.kwargs['username']})


def block_detail(request: HttpRequest, username, slug):
    """Detail view for block."""

    block = get_object_or_404(Block, user__username=username, slug=slug)

    # If the block is not private, show without authentication
    if block.private:
        if not request.user.is_authenticated:
            return redirect(reverse('cas_ng_login'))
        elif request.user.username != block.user.username:
            return redirect(reverse('index'))

    # POST request to upload and download ISP
    if request.method == 'POST':

        # If no supplied file, fall to default
        if not "excel_file" in request.FILES:
            template = ISPTemplate.objects.order_by('updated_at').first()
            if template is None:
                excel_file = os.path.join(settings.MEDIA_ROOT, 'isp_templates',
                                          'ISP_mall_default.xlsm')
            else:
                excel_file = os.path.join(settings.MEDIA_ROOT,
                                          template.template.name)
            wb = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=True)
        else:
            excel_file = request.FILES["excel_file"]
            # Check the file size and name length
            if (validator.file_validation(excel_file) != 0):
                return render(request, 'rodatraden/block/block_detail.html')

            wb = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=True)

        # regex is used later when matching course names. With this, only
        # alphabetic (swedish) characters and numbers will be used when comparing
        # course names.
        regex = re.compile('[^åäöÅÄÖa-zA-Z0-9]')

        # Create new list with block courses and credits, sort by course name
        block_courses = []
        for co in block.courseoccasions.all():
            block_courses.append([str(co.course.title), str(co.course.ects)])
        block_courses.sort(key = itemgetter(0))

        # loop through each sheet in excel
        for sheet in wb:

            # Specific sheet titles that contains courses
            if ((sheet.title == 'Profilkurser') |
                (sheet.title == 'Basterminer') |
                (sheet.title == 'Övriga kurser') |
                (sheet.title == 'Allmänna ingenjörskurser')):

                # loop through each course in excel, where:
                #   course[0]: Checkbox
                #   course[1]: Course name
                #   course[6]: Course credit
                for course in sheet.iter_rows(min_row=7, max_row=300,
                min_col=1, max_col=7):

                    # If there are no more courses on current sheet,
                    # go to next sheet
                    if not (course[1].value):
                        break

                    # loop through each course in the block list, where:
                    # index: Index of current element in ‘block_courses’
                    # item[0]: course name
                    # item[1]: course credit
                    for index, item in enumerate(block_courses):

                        # If a match of course name (case insensitive, disregard
                        # all whitespaces) and credit can be found, tick
                        # checkbox in excel and modify block list
                        if (regex.sub('', str(item[0]).lower())
                        == regex.sub('', str(course[1].value).lower())
                        and course[6].value
                        and math.isclose(float(item[1]),
                        float(course[6].value), abs_tol=0.1)):
                            course[0].value='x'
                            course[0].alignment = Alignment(horizontal
                            = "center", vertical = "bottom")
                            block_courses.pop(index)
                            break

        # Add all private courses to list
        # and add the list to a new sheet in excel
        for co in block.privatecourses.all():
            block_courses.append([str(co.title), str(co.ects)])

        if block_courses:
            count = 4
            for sheet in wb:
                if sheet.title == "Ej inlagda kurser":
                    wb.remove_sheet(sheet)
            ws = wb.create_sheet(title="Ej inlagda kurser")
            ws["A1"].value=("OBS! Lägg in dessa kurser manuellt under rätt "
            "flik. Töm denna kurslista innan du genererar sammanfattningen")
            ws["B3"].value='Kursnamn'
            ws["B3"].font=Font(bold=True)
            for item in block_courses:
                coursename_pos = 'B' + str(count)
                coursecredit_pos = 'C' + str(count)
                ws[coursename_pos].value = item[0]
                ws[coursecredit_pos].value = item[1]
                count = count + 1

        # Save workbook to memory buffer instead of file (no cleanup needed)
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Return the file as a download response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.ms-excel.sheet.macroEnabled.12'
        )
        response['Content-Disposition'] = f'attachment; filename="ISP_{block.slug}.xlsm"'
        return response
    else:
        # Get all categories for the block exam and build dict with those as
        # keys
        categories = CategoryExam.objects.filter(exam=block.exam)

        categories_title = [category.category.title for category in
            categories]

        categories_ects = [float(category.ects) for category in categories]

        category_sum = dict.fromkeys([category.category.title for category in
            categories], 0)
        # Get sum from block
        block.total_category_ects(category_sum)

        # XXX: The variables name will be confusingly similar until I figure out how category_sum works.
        categories_sum = [float(sum) for sum in category_sum.values()]

        course_occasions_json = [course.as_json() for course in block.courseoccasions.all()]
        prinvate_courses_json = [course.as_json() for course in block.privatecourses.all()]

        context = {
            'this_block': block,
            'start_year': block.start_year,
            'course_occasions': course_occasions_json + prinvate_courses_json,
            'categories_title': categories_title,
            'categories_ects': categories_ects,
            'categories_sum': categories_sum,
            'total_ects': block.total_course_ects(),
            'logged_in': (request.user.is_authenticated and
            request.user.username == block.user.username) or
            (request.user.is_staff and not block.private)
        }

        return render(request, 'rodatraden/block/block_detail.html', context)


class BlockImportList(CorrectUserPermissionMixin, LoginRequiredMixin, ListView):
    """List view for blocks that courses can be imported from."""

    model = Block
    template_name = 'rodatraden/block/block_import_list.html'

    def get_queryset(self, *args, **kwargs):
        qs = super().get_queryset()

        # Use can import form all public blocks published in a track and all
        # their own blocks.
        public_blocks = qs.filter(
            private=False
        ).exclude(
            track__isnull=True
        )
        private_blocks = qs.filter(user=self.request.user)
        all_blocks = public_blocks | private_blocks
        return all_blocks.order_by('title')
    

@login_required
def import_block(request: HttpRequest, username, slug):
    """Import course occasions from block from GET info."""

    # Block to import course occasions into.
    user_block = get_object_or_404(
        Block,
        user__username=username,
        slug=slug
    )

    # Block to import course occasions from.
    imported_block_username = request.GET.get('username', '')
    imported_block_slug = request.GET.get('slug', '')
    imported_block = get_object_or_404(
        Block,
        user__username=imported_block_username,
        slug=imported_block_slug
    )

    imported_course_occasions = import_course_occasions(
        user_block.start_year,
        imported_block
    )

    # Add imported course occasions to user block.
    for course_occasion in imported_course_occasions:
        try:
            user_block.courseoccasions.add(course_occasion)
        except:
            pass

    return redirect('block-detail', username=username, slug=slug)


@login_required
def block_course_list(request: HttpRequest, username: str, slug: str):
    """Custom list view with all courses that can be added a specific year and
    time period range."""

    # Get year and start week from GET request.
    try:
        year = int(request.GET.get('year'))
        start = int(request.GET.get('start'))
    except (TypeError, ValueError):
        raise Http404("Unspecified year or start week.")

    # The block is used for user validation and to know which courses have
    # already been added.
    block = get_object_or_404(Block, user__username=username, slug=slug)

    # Can only add courses to blocks the user owns, unless the user is staff.
    if not request.user.is_staff:
        if (block.user.username != request.user.username):
            return redirect(reverse('index'))

    # Get course IDs (not occasion IDs) that are already in the block anywhere.
    # This is used to mark courses as "already taking" (potential retakes).
    already_taking_course_ids = set(
        block.courseoccasions.values_list('course__id', flat=True)
    )

    # Get IDs of specific course occasions already in this block. Used to mark
    # them as "already added" so the user sees all options but knows which are
    # already in their schedule. This removes the previous restriction that
    # limited retakes by hiding already-added occasions.
    in_block_occasion_ids = set(
        block.courseoccasions.values_list('id', flat=True)
    )
    in_block_private_ids = set(
        block.privatecourses.values_list('id', flat=True)
    )

    # Get course IDs that START BEFORE this period (for prerequisite checking).
    # A course meets a prerequisite if it has started before the new course.
    courses_started_before = set()
    for co in block.courseoccasions.all():
        # Check if this course occasion starts before the period we're adding to
        if co.academic_year.year < year:
            courses_started_before.add(co.course.id)
        elif co.academic_year.year == year and co.time_period.week < start:
            courses_started_before.add(co.course.id)

    # Get ALL course occasions starting in the given period, including those
    # already in the block. Previously, already-in-block occasions were excluded
    # which effectively limited retakes. Now they are shown but marked as
    # "in_block" so the user can see all available options.
    courseoccasions = CourseOccasion.objects.filter(
        academic_year__year=year,
        time_period__week__gte=start,
        time_period__week__lt=start+10
    ).order_by('course__title')

    # Annotate each course occasion with retake and in-block status
    courseoccasions = list(courseoccasions)
    for co in courseoccasions:
        co.already_taking = co.course.id in already_taking_course_ids
        co.in_block = co.id in in_block_occasion_ids
        
        # Check if prerequisites are met
        co.has_unmet_prerequisites = False
        for prereq in co.course.prerequisites.all():
            # A prerequisite is met if ANY of the equivalent courses have started
            equivalent_ids = set(prereq.equivalent_courses.values_list('id', flat=True))
            if not equivalent_ids.intersection(courses_started_before):
                co.has_unmet_prerequisites = True
                break

    # Sort: in-block last, then not-already-taking first, then by title
    courseoccasions.sort(key=lambda co: (co.in_block, co.already_taking, co.course.title))

    # Get ALL private courses starting in the given period, including those
    # already in the block.
    privatecourses = PrivateCourse.objects.filter(
        user=request.user,
        year=year,
        start__gte=start,
        start__lt=start+10
    ).order_by('title')

    # Private courses don't have a shared course ID, so they can't be "retakes"
    # and don't have prerequisites
    privatecourses = list(privatecourses)
    for pc in privatecourses:
        pc.already_taking = False
        pc.has_unmet_prerequisites = False
        pc.in_block = pc.id in in_block_private_ids

    context = {
        'b_slug': slug,
        'username': username,
        'courseoccasions': courseoccasions,
        'privatecourses': privatecourses,
    }

    return render(request, 'rodatraden/block/block_course_list.html', context)


@login_required
def add_course_to_block(request: HttpRequest, block_username, block_slug):
    """Add a course to block from GET info."""

    course_occasion_slug = request.GET.get('slug', '')
    is_private = request.GET.get('private', '')

    block = get_object_or_404(Block, user__username=block_username, slug=block_slug)

    # Only admins are allowed to edit non-owned schedules.
    is_allowed_to_edit = (request.user.is_staff
                          or request.user.username == block.user.username)
    if not is_allowed_to_edit:
        raise PermissionDenied

    if (is_private == '1'):
        privatecourse = get_object_or_404(PrivateCourse, slug=course_occasion_slug)
        block.privatecourses.add(privatecourse)
    else:
        course = get_object_or_404(CourseOccasion, slug=course_occasion_slug)
        block.courseoccasions.add(course)

    # Return the updated set of private and non-private course occasions in the
    # block schedule as a single JSON array.
    course_occasions = block.courseoccasions.all()
    private_courses = block.privatecourses.all()
    course_occasions_json = [course.as_json() for course in course_occasions]
    private_courses_json = [course.as_json() for course in private_courses]
    # NOTE: "safe=False" is completely safe to use. See this answer for an
    # explaination:
    # https://stackoverflow.com/a/70204451/10844442
    return JsonResponse(course_occasions_json + private_courses_json,
                        safe=False)


@login_required
def remove_course_from_block(request: HttpRequest, block_username: str, block_slug: str):
    """Remove a course from a block."""

    course_occasion_slug = request.GET.get('slug', '')
    is_private = request.GET.get('private', '')

    block = get_object_or_404(Block, user__username=block_username, slug=block_slug)

    # Only admins are allowed to edit non-owned schedules.
    is_allowed_to_edit = (request.user.is_staff
                          or request.user.username == block.user.username)
    if not is_allowed_to_edit:
        raise PermissionDenied

    if (is_private == '1'):
        occasion = get_object_or_404(PrivateCourse, slug=course_occasion_slug)
        block.privatecourses.remove(occasion)
    else:
        occasion = get_object_or_404(CourseOccasion, slug=course_occasion_slug)
        block.courseoccasions.remove(occasion)

    # Return the updated set of private and non-private course occasions in the
    # block schedule as a single JSON array.
    course_occasions = block.courseoccasions.all()
    private_courses = block.privatecourses.all()
    course_occasions_json = [course.as_json() for course in course_occasions]
    private_courses_json = [course.as_json() for course in private_courses]
    # NOTE: "safe=False" is completely safe to use. See this answer for an
    # explaination:
    # https://stackoverflow.com/a/70204451/10844442
    return JsonResponse(course_occasions_json + private_courses_json,
                        safe=False)


@login_required
def replace_course_in_block(request: HttpRequest, block_username: str, block_slug: str):
    """Remove a course from a block and add another."""

    slug_to_remove = request.GET.get('slugToRemove', '')
    slug_to_add = request.GET.get('slugToAdd', '')

    block = get_object_or_404(Block, user__username=block_username, slug=block_slug)

    # Only admins are allowed to edit non-owned schedules.
    is_allowed_to_edit = (request.user.is_staff
                          or request.user.username == block.user.username)
    if not is_allowed_to_edit:
        raise PermissionDenied

    occasion_to_remove = get_object_or_404(CourseOccasion, slug=slug_to_remove)
    occasion_to_add = get_object_or_404(CourseOccasion, slug=slug_to_add)
    block.courseoccasions.remove(occasion_to_remove)
    block.courseoccasions.add(occasion_to_add)

    # Return the updated set of private and non-private course occasions in the
    # block schedule as a single JSON array.
    course_occasions = block.courseoccasions.all()
    private_courses = block.privatecourses.all()
    course_occasions_json = [course.as_json() for course in course_occasions]
    private_courses_json = [course.as_json() for course in private_courses]
    # NOTE: "safe=False" is completely safe to use. See this answer for an
    # explaination:
    # https://stackoverflow.com/a/70204451/10844442
    return JsonResponse(course_occasions_json + private_courses_json,
                        safe=False)


@login_required
def get_related_course_occasions(request: HttpRequest, block_username: str, block_slug: str):
    """Get all course occasions related to the same course as the specified
    course occasion."""

    course_occasion_slug = request.GET.get('slug', '')

    block = get_object_or_404(Block, user__username=block_username, slug=block_slug)

    occasion = get_object_or_404(CourseOccasion, slug=course_occasion_slug)
    course = occasion.course

    # Ignore the course occasion included in the request.
    course_occasions = course.courseoccasion_set.exclude(slug=course_occasion_slug)
    course_occasions_json = [occasion.as_json() for occasion in course_occasions]

    # NOTE: "safe=False" is completely safe to use. See this answer for an
    # explaination:
    # https://stackoverflow.com/a/70204451/10844442
    return JsonResponse(course_occasions_json, safe=False)


def get_block_category_sums(request: HttpRequest, block_username: str, block_slug: str):
    """Get the current category sums and total ECTS for a block schedule.
    
    This is used to update the category chart and total HP after adding/removing
    courses without requiring a full page reload.
    """

    block = get_object_or_404(Block, user__username=block_username, slug=block_slug)

    # Respect the same privacy rules as the block detail view.
    if block.private:
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        elif request.user.username != block.user.username:
            return JsonResponse({'error': 'Access denied'}, status=403)

    # Get all categories for the block exam
    categories = CategoryExam.objects.filter(exam=block.exam)

    # Build dict with category titles as keys
    category_sum = dict.fromkeys([category.category.title for category in
        categories], 0)
    
    # Get sum from block
    block.total_category_ects(category_sum)

    # Return the sums as a list and the total ECTS
    categories_sum = [float(sum) for sum in category_sum.values()]
    total_ects = block.total_course_ects()

    return JsonResponse({
        'categorySums': categories_sum,
        'totalEcts': float(total_ects)
    })


@login_required
def update_prerequisite_check(request: HttpRequest, username, slug):
    """Update if the prerequisites should be verified for the block or not."""

    block = get_object_or_404(Block, user__username=username, slug=slug)

    shouldEnable = request.GET.get('enable', '')

    # Treat anything non-empty as True
    # XXX: Will also be true for "enable=0" or "enable=false" etc.
    if (shouldEnable):
        block.should_verify_prerequisites = True
    else:
        block.should_verify_prerequisites = False

    block.save()

    # Just return an empty response if everything went well.
    return HttpResponse('')
